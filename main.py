import openpyxl
from datetime import datetime, timedelta
from PyQt6.QtCore import Qt
from PyQt6.QtWidgets import *
from PyQt6 import QtGui
from gui import Ui_Dialog
import shutil
import os
import sys
import re
import pandas as pd


def date_plus_day_to_str(date_obj, days = 1, str_format = "%Y년 %m월 %d일"):
    after_oneday_obj = date_obj + timedelta(days=days)
    formatted_date = after_oneday_obj.strftime(str_format)
    return formatted_date


def is_xlsx(file_path):
    # 파일이고 확장자가 .xlsx인지 확인
    return os.path.isfile(file_path) and file_path.lower().endswith('.xlsx')


class WindowClass(QMainWindow, Ui_Dialog):
    def __init__(self):
        super().__init__()
        self.ui = Ui_Dialog()
        self.setupUi(self)
        self.setAcceptDrops(True)
        self.setWindowTitle("Auto Coupang report")

        self.setFixedWidth(915)
        self.setFixedHeight(554)

        font1 = QtGui.QFont("새굴림", 15)
        h1_list = [self.label_1, self.label_2, self.label_3, self.label_4]
        [obj.setFont(font1) for obj in h1_list]

        self.input_file_path_1 = ""
        self.input_file_path_2 = ""
        self.input_file_path_3 = ""
        self.output_file_path = ""

        self.btn_search_excel_1.clicked.connect(self.browse_1)
        self.btn_search_excel_2.clicked.connect(self.browse_2)
        self.btn_search_excel_3.clicked.connect(self.browse_3)
        self.btn_execute.clicked.connect(self.push_execute)
        self.gui_home_dir = os.path.abspath(".")
        self.is_successed = True

    def browse_1(self):
        try:
            file_filter = "Excel Files (*.xlsx)"
            path, _ = QFileDialog.getOpenFileName(None, 'Select an Excel File', '', file_filter)
            self.input_file_path_1 = path
        except Exception as e:
            print(e)
            self.input_file_path_1 = ""
        self.textEdit_log.setText(f"{self.textEdit_log.toPlainText()}쿠팡 일일 데이터 엑셀 위치: {self.input_file_path_1}\n")
        self.lineEdit_excel_1.setText(f"{self.input_file_path_1}")


    def browse_2(self):
        try:
            file_filter = "Excel Files (*.xlsx)"
            path, _ = QFileDialog.getOpenFileName(None, 'Select an Excel File', '', file_filter)
            self.input_file_path_2 = path
        except:
            self.input_file_path_2 = ""
        self.textEdit_log.setText(f"{self.textEdit_log.toPlainText()}쿠팡 누적 데이터 엑셀 위치: {self.input_file_path_2}\n")
        self.lineEdit_excel_2.setText(f"{self.input_file_path_2}")


    def browse_3(self):
        try:
            file_filter = "Excel Files (*.xlsx)"
            path, _ = QFileDialog.getOpenFileName(None, 'Select an Excel File', '', file_filter)
            self.input_file_path_3 = path
        except:
            self.input_file_path_3 = ""
        self.textEdit_log.setText(f"{self.textEdit_log.toPlainText()}적용할 리포트 엑셀 위치: {self.input_file_path_3}\n")
        self.lineEdit_excel_3.setText(f"{self.input_file_path_3}")


    def finish_task(self, rst):
        self.btn_search_excel_1.setEnabled(True)
        self.btn_search_excel_2.setEnabled(True)
        self.btn_search_excel_3.setEnabled(True)
        if rst and self.is_successed:
            self.textEdit_log.setText(f"{self.textEdit_log.toPlainText()}작업이 완료 됐습니다.\n생성한 리포트 파일 위치: {self.output_file_path}\n")
        else:
            self.textEdit_log.setText(f"{self.textEdit_log.toPlainText()}{self.input_file_path_1}sheet가 여러개 입니다. 실행 종료\n")
        self.is_successed = True


    def process_file_path(self, report_day):
        pattern = r"\/\d{6}_"  # 6자리 숫자 (YYMMDD) 패턴
        match = re.search(pattern, self.input_file_path_3)

        if match:
            old_date = match.group()
            new_date = report_day.strftime("%y%m%d")
            self.output_file_path = self.input_file_path_3.replace(old_date, f"/{new_date}_")


    def auto_report(self):
        # data_xls_name = "data/20241105_data.xlsx"
        # report_template_xls_name = "data/241106_dearcamp.xlsx"
        workbook_dayily = openpyxl.load_workbook(self.input_file_path_1)
        sheet = workbook_dayily.active  # 기본 활성화된 시트를 선택합니다. 특정 시트가 있으면 sheet = workbook["Sheet1"] 과 같이 지정합니다.
        sheets = workbook_dayily.sheetnames
        ## Excel 데이터 to 리스트
        ### Daily 
        row_values_daily = []
        for row_idx, row in enumerate(sheet):
            if row_idx == 0:
                continue
            row_data = [col.value for col in row]
            row_values_daily.append(row_data)
        workbook_dayily.close()
        ### total 
        workbook_cum = openpyxl.load_workbook(self.input_file_path_2)
        sheet = workbook_cum.active  # 기본 활성화된 시트를 선택합니다. 특정 시트가 있으면 sheet = workbook["Sheet1"] 과 같이 지정합니다.
        sheets = workbook_cum.sheetnames
        row_values_cumul = []
        for row_idx, row in enumerate(sheet):
            if row_idx == 0:
                continue
            row_data = [col.value for col in row]
            row_values_cumul.append(row_data)
        workbook_cum.close()
        ## 리스트 to Data
        ### Daily
        dataset_daily = {}
        for row in row_values_daily:
            date_obj = datetime.strptime(str(int(row[0])), "%Y%m%d").date()
            temp = {}
            temp["campaign"] = row[6]
            temp["exposure_num"] = row[7]
            temp["click_num"] = row[8]
            temp["ad_cost"] = row[9]
            temp["convert_num"] = row[15]
            temp["convert_cost"] = row[18]
            try:
                dataset_daily[date_obj].append(temp)
            except KeyError:
                dataset_daily[date_obj] = [temp]
        ### total 
        dataset_cumul = []
        for row in row_values_cumul:
            temp = {}
            temp["campaign_name"] = row[4]
            if row[11]:
                keyword = row[11]
            else:
                keyword = "비검색 영역"
            temp["keyword"] = keyword
            temp["exposure_num"] = row[12]
            temp["click_num"] = row[13]
            temp["ad_cost"] = row[14]
            temp["convert_num"] = row[28]
            temp["convert_cost"] = row[31]
            dataset_cumul.append(temp)
        df = pd.DataFrame(dataset_cumul)
        result_campaign = df.groupby("campaign_name")[["exposure_num", "click_num", "ad_cost", "convert_num", "convert_cost"]].sum()
        result_by_keyword = df.groupby("keyword")[["exposure_num", "click_num", "ad_cost", "convert_num", "convert_cost"]].sum()
        result_by_keyword = result_by_keyword.sort_values(by=["ad_cost", "convert_num"], ascending=[False, False]).reset_index()

        if len(sheets) != 1:
            return False
        report_day = max(dataset_daily.keys())
        file_day = report_day + timedelta(days=1)
        self.process_file_path(file_day)
        try:
            shutil.copy(self.input_file_path_3, self.output_file_path)
        except PermissionError:
            self.textEdit_log.setText(f"{self.textEdit_log.toPlainText()}{self.output_file_path}파일이 이미 존재하거나 열려있어 실패.\n")
            self.is_successed = False
        #################################################### 리포트 파일
        #################################################### ["요약"] 시트
        report_wb = openpyxl.load_workbook(self.output_file_path)
        report_sheet = report_wb["요약"]
        date_str = date_plus_day_to_str(report_day, 1, "%Y년 %m월 %d일")
        report_sheet["B31"] = date_str
        date_str = date_plus_day_to_str(report_day, 1, "%Y-%m-%d")
        report_sheet["C104"] = date_str
        #################################################### ["쿠팡_일일"] 시트
        report_sheet = report_wb["쿠팡_일일"]
        for date_obj in dataset_daily.keys():
            for row_idx, cell in enumerate(report_sheet["B"]):
                try:
                    cv_date = cell.value.date()
                except AttributeError:
                    continue
                if cv_date.strftime("%y-%m-%d") == date_obj.strftime("%y-%m-%d"):
                    data_idx = row_idx + 1
                    report_sheet[f"D{data_idx}"] = sum(campaign["exposure_num"] for campaign in dataset_daily[date_obj])
                    report_sheet[f"E{data_idx}"] = sum(campaign["click_num"] for campaign in dataset_daily[date_obj])
                    report_sheet[f"H{data_idx}"] = sum(campaign["ad_cost"] for campaign in dataset_daily[date_obj])
                    report_sheet[f"I{data_idx}"] = sum(campaign["convert_num"] for campaign in dataset_daily[date_obj])
                    report_sheet[f"L{data_idx}"] = sum(campaign["convert_cost"] for campaign in dataset_daily[date_obj])
                    continue
        #################################################### ["쿠팡_누적"] 시트
        report_sheet = report_wb["쿠팡_누적"]
        key_col_dict_c = {"campaign_name": "C", "exposure_num": "D", "click_num": "E", "ad_cost": "H", "convert_num": "I", "convert_cost": "L"}
        rows = [9, 10]

        for row in rows:
            # Get the campaign name from the Excel sheet
            campaign = report_sheet[f"{key_col_dict_c['campaign_name']}{row}"].value
            
            if campaign in result_campaign.index:
                for key, col in key_col_dict_c.items():
                    if key != "campaign_name":  # Skip writing the campaign name back
                        # Populate the cell with the corresponding value from the DataFrame
                        report_sheet[f"{col}{row}"] = result_campaign.loc[campaign, key]

        key_col_dict_k = {"keyword": "C", "exposure_num": "D", "click_num": "E", "ad_cost": "H", "convert_num": "I", "convert_cost": "L"}
        start_row = 33
        for idx, row in result_by_keyword.iterrows():
            for key, col in key_col_dict_k.items():
                report_sheet[f"{col}{start_row+idx}"] = row[key]

        # 변경사항 저장
        report_wb.save(self.output_file_path)
        report_wb.close()
        return True

    def push_execute(self):
        rst = self.auto_report()
        self.finish_task(rst)


if __name__ == "__main__":
    app = QApplication(sys.argv)
    myWindow = WindowClass()
    myWindow.show()
    app.exec()